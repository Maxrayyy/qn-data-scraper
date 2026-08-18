from pathlib import Path

from openpyxl import load_workbook

from app.scraper import ProductItem
from app.worker import export_partial_rows, is_our_browser_proc


def _collect_log(entries: list):
    def log(level: str, msg: str) -> None:
        entries.append((level, msg))

    return log


def test_is_our_browser_true_for_bundled_chromium():
    assert is_our_browser_proc(
        "chrome.exe",
        r"C:\app\千牛数据抓取工具\_internal\browsers\chromium-1148\chrome-win\chrome.exe",
        Path(r"C:\app\千牛数据抓取工具\_internal"),
    )


def test_is_our_browser_false_for_user_chrome():
    assert not is_our_browser_proc(
        "chrome.exe",
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        Path(r"C:\app\千牛数据抓取工具\_internal"),
    )


def test_is_our_browser_false_for_non_browser():
    assert not is_our_browser_proc("python.exe", r"C:\app\_internal\python.exe", Path(r"C:\app\_internal"))


def test_export_partial_rows_writes_excel(tmp_path):
    sink = [
        (ProductItem(name="商品A", product_id="1001"), None),
        (ProductItem(name="商品B", product_id="1002"), None),
    ]
    entries: list = []
    n = export_partial_rows(sink, str(tmp_path), _collect_log(entries))
    assert n == 2
    out = tmp_path / "千牛商品发现数据.xlsx"
    assert out.exists()
    wb = load_workbook(out)
    ws = wb.active
    assert ws.max_row == 3  # 表头 + 2 行数据
    assert any("已导出" in msg for _, msg in entries)


def test_export_partial_rows_empty_no_file(tmp_path):
    entries: list = []
    n = export_partial_rows([], str(tmp_path), _collect_log(entries))
    assert n == 0
    assert list(tmp_path.iterdir()) == []


# ---------- 持久化浏览器会话（像正常浏览器一样记住登录状态） ----------

def test_persistent_context_smoke(tmp_path):
    """launch_persistent_context 是产品依赖的启动方式，验证其可用性。"""
    from playwright.sync_api import sync_playwright

    with sync_playwright() as p:
        ctx = p.chromium.launch_persistent_context(
            user_data_dir=str(tmp_path / "prof"), headless=True
        )
        try:
            page = ctx.pages[0] if ctx.pages else ctx.new_page()
            page.set_content("<div>ok</div>")
            assert "ok" in page.inner_text("body")
        finally:
            ctx.close()
