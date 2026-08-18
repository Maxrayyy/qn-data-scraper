import base64
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from app.excel_export import (
    HEADERS,
    build_output_path,
    dedupe_rows,
    export_to_excel,
)
from app.scraper import ProductItem

# 1x1 透明 PNG（硬编码，避免引入 Pillow）
PNG_1PX = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNkYPhfDwAChwGA60e6kgAAAABJRU5ErkJggg=="
)


def _make_item(pid="123456"):
    return ProductItem(
        image_url="https://img.alicdn.com/x.jpg",
        name="猫别墅A",
        item_url=f"https://item.taobao.com/item.htm?id={pid}",
        product_id=pid,
        orders="1,234",
        price="¥99.00",
    )


def test_build_output_path_no_collision(tmp_path):
    p = build_output_path(tmp_path)
    assert p.name == "千牛商品发现数据.xlsx"


def test_build_output_path_collision_append_timestamp(tmp_path):
    (tmp_path / "千牛商品发现数据.xlsx").write_bytes(b"old")
    p = build_output_path(tmp_path, now=datetime(2026, 8, 18, 15, 30, 12))
    assert p.name == "千牛商品发现数据_20260818_153012.xlsx"
    assert p != tmp_path / "千牛商品发现数据.xlsx"  # 不覆盖旧文件


def test_export_headers_values_hyperlink_image(tmp_path):
    img_path = tmp_path / "img.png"
    img_path.write_bytes(PNG_1PX)
    out = export_to_excel(
        [(_make_item(), img_path), (_make_item("999999"), None)],
        tmp_path / "out.xlsx",
    )
    assert out.exists()
    wb = load_workbook(out)
    ws = wb.active
    assert [ws.cell(row=1, column=i).value for i in range(1, 6)] == HEADERS
    assert ws.cell(row=2, column=3).value == "123456"
    assert ws.cell(row=2, column=4).value == "1,234"
    assert ws.cell(row=2, column=5).value == "¥99.00"
    assert ws.cell(row=2, column=2).hyperlink.target.endswith("id=123456")
    assert len(ws._images) == 1          # 第一行有图片，第二行无
    assert ws.row_dimensions[2].height == 64


# ---------- 坏图降级 ----------

def test_export_corrupt_image_degrades(tmp_path):
    # CDN 把 200 HTML 错误页存成 .jpg：PIL 打不开 → 本行仅保留文字，导出不中断
    img_path = tmp_path / "bad.jpg"
    img_path.write_bytes(b"not an image")
    out = export_to_excel([(_make_item(), img_path)], tmp_path / "out.xlsx")
    assert out.exists()
    wb = load_workbook(out)
    ws = wb.active
    assert len(ws._images) == 0
    assert ws.cell(row=2, column=2).value == "猫别墅A"
    assert ws.cell(row=2, column=3).value == "123456"


# ---------- 导出去重（按 product_id / 三元组） ----------

def _make_no_id_item(name="无名商品", orders="1,234", price="¥99.00"):
    return ProductItem(name=name, orders=orders, price=price, product_id="")


def test_dedupe_rows_by_product_id():
    rows = dedupe_rows([(_make_item("123456"), None), (_make_item("123456"), None)])
    assert len(rows) == 1
    assert rows[0][0].product_id == "123456"


def test_export_dedupes_same_product_id(tmp_path):
    out = export_to_excel(
        [(_make_item("123456"), None), (_make_item("123456"), None)],
        tmp_path / "out.xlsx",
    )
    wb = load_workbook(out)
    ws = wb.active
    assert ws.max_row == 2  # 表头 + 1 行数据


def test_export_dedupes_empty_id_same_triple(tmp_path):
    out = export_to_excel(
        [(_make_no_id_item(), None), (_make_no_id_item(), None)],
        tmp_path / "out.xlsx",
    )
    wb = load_workbook(out)
    ws = wb.active
    assert ws.max_row == 2


def test_export_keeps_empty_id_different_triple(tmp_path):
    out = export_to_excel(
        [
            (_make_no_id_item(name="商品A", orders="1", price="¥1.00"), None),
            (_make_no_id_item(name="商品B", orders="2", price="¥2.00"), None),
        ],
        tmp_path / "out.xlsx",
    )
    wb = load_workbook(out)
    ws = wb.active
    assert ws.max_row == 3  # 表头 + 2 行数据
