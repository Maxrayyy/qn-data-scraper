"""Excel 导出：真实图片嵌入 + 商品名超链接 + 时间戳防覆盖。"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill

from .scraper import ProductItem

HEADERS = ["商品图片", "商品名", "商品id", "支付单量", "件单价"]
BASE_FILENAME = "千牛商品发现数据"
IMAGE_HEIGHT_PX = 60  # 嵌入图片统一缩略高度（像素）
ROW_HEIGHT = 64       # 数据行高


def dedupe_rows(rows: list[tuple[ProductItem, Path | None]]) -> list[tuple[ProductItem, Path | None]]:
    """按 product_id 去重（保留首次出现）；product_id 为空时按 (商品名, 支付单量, 件单价) 三元组去重。"""
    seen: set[tuple] = set()
    out: list[tuple[ProductItem, Path | None]] = []
    for row in rows:
        item = row[0]
        key = (
            (item.product_id,)
            if item.product_id
            else (item.name, item.orders, item.price)
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(row)
    return out


def build_output_path(export_dir: str | Path, now: datetime | None = None) -> Path:
    """文件名冲突时自动追加时间戳，绝不覆盖旧文件。now 参数仅供测试注入。"""
    now = now or datetime.now()
    d = Path(export_dir)
    path = d / f"{BASE_FILENAME}.xlsx"
    if not path.exists():
        return path
    return d / f"{BASE_FILENAME}_{now:%Y%m%d_%H%M%S}.xlsx"


def export_to_excel(
    rows: list[tuple[ProductItem, Path | None]], output_path: Path
) -> Path:
    """rows: (商品数据, 已下载的本地图片路径；下载失败为 None)。

    多条"商品发现"的数据直接依次追加到同一个工作簿。
    """
    rows = dedupe_rows(rows)  # 验证码重跑/关弹窗失败可能重复抓取，导出前先去重
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "商品发现数据"
    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    for col, title in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    # 数据行
    for i, (item, img_path) in enumerate(rows, start=2):
        ws.row_dimensions[i].height = ROW_HEIGHT
        # 商品图片：嵌入真实图片（缩略到统一高度，保持宽高比）。
        # CDN 返回 200 的 HTML 错误页被存成 .jpg 时 PIL 打不开 → 本行降级为仅保留文字。
        if img_path and img_path.exists():
            try:
                img = XLImage(str(img_path))
                img.width = IMAGE_HEIGHT_PX * (img.width / img.height)
                img.height = IMAGE_HEIGHT_PX
                ws.add_image(img, f"A{i}")
            except Exception:
                pass  # 单张坏图不中断整个导出
        # 商品名：带超链接
        name_cell = ws.cell(row=i, column=2, value=item.name or "(无名称)")
        if item.item_url:
            name_cell.hyperlink = item.item_url
            name_cell.style = "Hyperlink"
        ws.cell(row=i, column=3, value=item.product_id)
        ws.cell(row=i, column=4, value=item.orders)
        ws.cell(row=i, column=5, value=item.price)
        for col in range(1, 6):
            ws.cell(row=i, column=col).alignment = Alignment(vertical="center")
    # 列宽
    for col, width in zip("ABCDE", (16, 46, 20, 12, 12)):
        ws.column_dimensions[col].width = width
    wb.save(output_path)
    return output_path
